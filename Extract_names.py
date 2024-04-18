import requests
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# ! Global variables to modify before running the script !
api_key = "XXXXX" # Get this from your OpenSea profile => Settings => Developer

# Please note that this script works until meme card #400. After that, you'll need to implement a loop to iterate through the "next" string

def fetch_nft_data(collection_slug):
    # First URL with an empty 'next' parameter
    url = f"https://api.opensea.io/api/v2/collection/{collection_slug}/nfts?limit=200&next="
    headers = {
        "Accept": "application/json",
        "X-API-KEY": api_key
    }
    
    # Make the first request
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print(f"Failed to fetch NFT data: {response.status_code}")
        return None
    
    # Extract data from the first request
    data = response.json()
    all_nfts = data.get('nfts', [])
    
    # Check if there is a 'next' page
    next_page = data.get('next', None)
    
    # If there is a next page, make a second request
    if next_page:
        next_url = f"{url}{next_page}"  # Append the 'next' token to the URL
        response = requests.get(next_url, headers=headers)
        if response.status_code == 200:
            more_data = response.json()
            all_nfts += more_data.get('nfts', [])  # Append additional NFTs from the next page
        else:
            print(f"Failed to fetch additional NFT data: {response.status_code}")

    return {'nfts': all_nfts}  # Return a dictionary with 'nfts' as key

def update_nft_sheet(filename, sheet_name, new_data):
    # Load the workbook and select the specified sheet, or create it if it doesn't exist
    book = openpyxl.load_workbook(filename)
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        # Clear existing entries starting from row 2
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
    else:
        # Create new sheet and add headers
        sheet = book.create_sheet(sheet_name)
        sheet.append(['Identifier', 'Name'])

    # Append new rows from the new data
    for nft in new_data:
        sheet.append([nft['Identifier'], nft['Name']])
        # Set number format for the Identifier as integer
        sheet.cell(row=sheet.max_row, column=1).number_format = '0'

    # Save the updated workbook
    book.save(filename)
    print("Updated the Excel file with new NFT data.")


def main():
 
    collection_slug = "thememes6529"
    data = fetch_nft_data(collection_slug)
    
    if data and 'nfts' in data:
        nft_list = [{'Identifier': int(nft['identifier']), 'Name': nft['name']} for nft in data['nfts']]
        update_nft_sheet('listings.xlsx', 'Cards_names', nft_list)
    else:
        print("No data found or failed to fetch data.")

if __name__ == "__main__":
    main()
