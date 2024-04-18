import os
import requests
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# ! Global variables to modify before running the script !
api_key = "XXXXX" # Get this from your OpenSea profile => Settings => Developer
nb_memecards = 219 # how many meme cards there are in the collection

def update_excel_file(filename, sheet_name, df):
    # Load the existing workbook or create a new one
    if os.path.exists(filename):
        book = openpyxl.load_workbook(filename)
    else:
        book = openpyxl.Workbook()
        book.save(filename)

    # Check if the specified sheet exists; if not, create it
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(sheet_name)

    # Erase the contents of columns A to F from the second row onwards
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

    # Write new data, including headers
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, start=1):  # Start from row 1 to include headers
        for c_idx, value in enumerate(row, start=1):  # Start from column 1
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Save the modified workbook
    book.save(filename)
    print("Excel file has been updated.")

def fetch_best_listing(collection_slug, item_id):
    url = f"https://api.opensea.io/api/v2/listings/collection/{collection_slug}/nfts/{item_id}/best"
    headers = {
        "Accept": "application/json",
        "X-API-KEY": api_key
    }
    response = requests.get(url, headers=headers)
    return response.json() if response.status_code == 200 else None

def fetch_best_bid(collection_slug, item_id):
    url = f"https://api.opensea.io/api/v2/offers/collection/{collection_slug}/nfts/{item_id}/best"
    headers = {
        "Accept": "application/json",
        "X-API-KEY": api_key
    }
    response = requests.get(url, headers=headers)
    return response.json() if response.status_code == 200 else None

def main():
    collection_slug = "thememes6529"
    items = range(1, nb_memecards + 1)  # From item 1 to how many cards defined in global variable
    
    data = []
    for item_id in items:
        listing = fetch_best_listing(collection_slug, item_id)
        bid = fetch_best_bid(collection_slug, item_id)
        
        price_eth = int(listing['price']['current']['value']) / 10**18 if listing else 'N/A'
        bid_eth = int(bid['price']['value']) / 10**18 if bid else 'N/A'
        spread_eth = abs(price_eth - bid_eth) if isinstance(price_eth, float) and isinstance(bid_eth, float) else 'N/A'
        spread_percentage = (spread_eth / price_eth) * 100 if price_eth != 0 else 'N/A'

        item_data = {
            'Item ID': item_id,
            'Order Hash': listing.get('order_hash', 'N/A') if listing else 'N/A',
            'Price (ETH)': price_eth,
            'Best Bid (ETH)': bid_eth,
            'Spread (ETH)': spread_eth,
            'Spread (%)': spread_percentage
        }
        data.append(item_data)
    
    df = pd.DataFrame(data)
    # Update or create Excel file and specified sheet
    update_excel_file('listings.xlsx', 'listings_bids', df)
    print("Data has been written to 'listings.xlsx'.")

if __name__ == "__main__":
    main()
