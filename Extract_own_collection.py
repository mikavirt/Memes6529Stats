import requests
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# ! Global variables to modify before running the script !
api_key = "xxxxxx" # Get this from your OpenSea profile => Settings => Developer
wallet_addresses = ["0xAAAAA", "0xBBBB", "0xCCCC"]
    

def fetch_nfts_owned_by_wallet(collection_slug, limit=120):
    """Fetch NFTs owned by specific wallets for a given collection."""
    nfts = []
    for wallet_address in wallet_addresses:
        url = f"https://api.opensea.io/api/v2/chain/ethereum/account/{wallet_address}/nfts?collection={collection_slug}&limit={limit}"
        headers = {
            "Accept": "application/json",
            "X-API-KEY": api_key
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            nft_data = response.json()
            if nft_data and 'nfts' in nft_data:
                nfts.extend([{'Identifier': int(nft['identifier']), 'Name': nft['name'], 'Quantity': 1, 'Wallet Address': wallet_address, 'OpenSea URL': nft['opensea_url']} for nft in nft_data['nfts']])
        else:
            print(f"Failed to fetch NFT data for wallet {wallet_address}: {response.status_code}")
    return nfts

def update_excel_sheet(filename, sheet_name, new_data):
    # Load or create the workbook
    if os.path.exists(filename):
        book = openpyxl.load_workbook(filename)
    else:
        book = openpyxl.Workbook()

    # Load or create the sheet
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(sheet_name)
        sheet.append(['Identifier', 'Name', 'Quantity', 'Wallet Address', 'OpenSea URL'])  # Add headers if new sheet

    # Clear existing data starting from the second row
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

    # Append new data
    df = pd.DataFrame(new_data)
    df['Identifier'] = pd.to_numeric(df['Identifier'], errors='coerce')  # Ensure identifiers are numeric for sorting
    df.sort_values(by='Identifier', ascending=False, inplace=True)  # Sort ascending for easier consecutive checking

    for _, row in df.iterrows():
        sheet.append([str(row['Identifier']), row['Name'], row['Quantity'], row['Wallet Address'], row['OpenSea URL']])

    # Apply highlighting to consecutive duplicates
    green_fill = PatternFill(start_color="E0FFCC", end_color="E0FFCC", fill_type="solid")
    previous_id = None
    previous_row = None

    for row in range(2, sheet.max_row + 1):  # Start from the first data row
        current_id = sheet.cell(row=row, column=1).value  # Get the identifier of the current row
        if current_id == previous_id:
            # If current and previous identifiers are the same, highlight both rows
            for col in range(1, 6):  # Assuming there are 5 columns
                sheet.cell(row=row, column=col).fill = green_fill
                if previous_row:
                    sheet.cell(row=previous_row, column=col).fill = green_fill
        previous_id = current_id
        previous_row = row
        
    book.save(filename)
    print(f"Updated Excel file '{filename}' in sheet '{sheet_name}'.")

def main():
    collection_slug = "thememes6529"
    
    nfts = fetch_nfts_owned_by_wallet(collection_slug)
    
    if nfts:
        update_excel_sheet('listings.xlsx', 'Collection', nfts)
    else:
        print("No NFT data found or failed to fetch data.")

if __name__ == "__main__":
    main()
